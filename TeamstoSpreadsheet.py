# script to read csv file and translate it to spreadsheet, probably excel....

from openpyxl import load_workbook

file = r'C:\Users\Wayne\Documents\WorkdayUpdatesandResources.xlsx'
wb_members = load_workbook(file)
num_to_col = {
    1: 'B',
    2: 'C',
    3: 'D',
    4: 'E',
    5: 'F',
}
Organization = [
    'University of Arkansas System Division of Agriculture',
    'University of Arkansas Grantham',
    'University of Arkansas forMedical Sciences',
    'University of Arkansas for Medical Sciences',
    'University of Arkansas atPine Bluff',
    'University of Arkansas at Pine Bluff',
    'University of Arkansas atMonticello',
    'University of Arkansas at Monticello',
    'University of Arkansas',
    'University Arkansas Community College Morrilton',
    'UAHT.EDU',
    'UACCB',
    'UA Rich Mountain',
    'UA Little Rock',
    'UA Cossatot',
    'UA - Pulaski Technical College',
    'Phillips Community College',
    'ASMSA',
    'University of Arkansas System Office',
]
sheet_names = [
    'Change Management',
    'Comms and Resources',
    'Institution Toolkits',
    'Problem Management',
    'Release Management',
    'SharePoint Resources',
]

for sheet in sheet_names:
    ws_change_management = wb_members[sheet]
    for cell in ws_change_management['A']:
        cell_data = cell.value
        cell_next = int(cell.row) + 1
        # don't know what field for what ???
        check = int()
        if cell_data is not None:
            print(cell.coordinate, 'is not None.')
            print('Then how many to next None?')
            # this is how many cells are next before there is an empty cell
            until_empty = int(0)
            for row in ws_change_management.iter_rows(min_row=cell_next, max_col=cell.column, values_only=True):
                print(row[0])
                if row[0] is not None:
                    until_empty += 1
                else:
                    break
            print('This is how many rows are before the next empty row: ', until_empty)
            i = 1
            while i <= until_empty:
                # figuring out how to manipulate data with until_empty
                row_number = int(cell.row) + i
                # moving data from one column and putting it into B and removing the data from A
                data = ws_change_management['A' + str(row_number)].value
                column = str(num_to_col[i])
                print('This is the column to use: ', column)
                for box in ws_change_management[column]:
                    if column == 'B':
                        check = box.row
                    if box.value is None and column == 'B':
                        ws_change_management[str(num_to_col[i]) + str(box.row)].value = data
                        ws_change_management['A' + str(row_number)].value = None
                        wb_members.save(file)
                        break
                    elif box.row == check and box.value is None:
                        ws_change_management[str(num_to_col[i]) + str(box.row)].value = data
                        ws_change_management['A' + str(row_number)].value = None
                        wb_members.save(file)
                        break
                i += 1
    # moving the name up in the field
    for box in ws_change_management['A']:
        if box.value is None:
            for row in ws_change_management.iter_rows(min_row=box.row, max_col=box.column):
                if row[0].value is not None:
                    next_name_found = ws_change_management['A' + str(row[0].row)].value
                    ws_change_management['A' + str(box.row)].value = next_name_found
                    ws_change_management['A' + str(row[0].row)].value = None
                    wb_members.save(file)
                    break

    # going to write a script that sorts the columns data to their respective column
    for row in ws_change_management.iter_rows(min_row=1, min_col=1, max_col=5):
        i = 0
        while i < 4:
            if row[i].value == 'Member':
                ws_change_management[str(num_to_col[i + 1]) + str(row[i].row)].value = 'N/A'
                ws_change_management['E' + str(row[i].row)].value = 'Member'
            i += 1
    wb_members.save(file)
    for cell in ws_change_management['D']:
        word = str(cell.value)
        if word == 'Member':
            ws_change_management['D' + str(cell.row)].value = 'N/A'
    wb_members.save(file)
    for row in ws_change_management.iter_rows(min_row=1, min_col=1, max_col=5):
        i = 0
        while i < 4:
            if row[i].value is None:
                ws_change_management[str(num_to_col[i + 1]) + str(row[i].row)].value = 'N/A'
            i += 1
    # check columns B and move organization to C if found
    for cell in ws_change_management['B']:
        word = str(cell.value)
        for match in Organization:
            if match == word:
                data_in_c = ws_change_management['C' + str(cell.row)].value
                if data_in_c == 'N/A':
                    ws_change_management['C' + str(cell.row)].value = match
                    ws_change_management['B' + str(cell.row)].value = 'N/A'
                else:
                    ws_change_management['D' + str(cell.row)].value = data_in_c
                    ws_change_management['C' + str(cell.row)].value = match
                    ws_change_management['B' + str(cell.row)].value = 'N/A'
    wb_members.save(file)
